"""
Анализ типов дисплеев и поиск моделей с несколькими вариантами
"""
import openpyxl
import re
from collections import defaultdict

wb = openpyxl.load_workbook(r"C:\Users\user\Desktop\Дисплеи.xlsx", data_only=True)
ws = wb.active

# Собираем все данные
all_data = []
for row in range(1, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value or ""
    price = ws.cell(row=row, column=2).value
    all_data.append((name, price))

# Извлекаем тип дисплея из названия
def extract_display_type(name):
    name_upper = name.upper()
    # Порядок проверки важен - сначала более специфичные
    if 'IN-CELL' in name_upper:
        return 'In-Cell'
    if 'AMOLED' in name_upper:
        return 'AMOLED'
    if 'OLED' in name_upper:
        return 'OLED'
    if ' OR ' in name or '- OR' in name:
        return 'OR'  # Оригинальный
    if 'СТАНДАРТ' in name_upper:
        return 'Стандарт'
    if 'LCD' in name_upper:
        return 'LCD'
    if 'TFT' in name_upper:
        return 'TFT'
    # Если нет типа - это базовая версия
    return 'Стандарт'

# Убираем тип дисплея чтобы получить базовую модель
def get_base_model(name):
    # Убираем "Дисплей "
    model = name.replace("Дисплей ", "").strip()
    # Убираем тип дисплея
    model = re.sub(r'\s*-\s*(OR|OLED|AMOLED|IN-CELL|LCD|TFT|Стандарт)\b', '', model, flags=re.IGNORECASE)
    # Убираем примечания в скобках про раму
    model = re.sub(r'\s*\(.*?рама.*?\)', '', model)
    model = re.sub(r'\s*\(отпечаток работает\)', '', model)
    return model.strip()

# Группируем по базовой модели
model_groups = defaultdict(list)
for name, price in all_data:
    base = get_base_model(name)
    dtype = extract_display_type(name)
    model_groups[base].append({
        'full_name': name,
        'price': price,
        'type': dtype
    })

# Находим модели с несколькими вариантами
multi_variant = {k: v for k, v in model_groups.items() if len(v) > 1}

print(f"Всего записей: {len(all_data)}")
print(f"Уникальных базовых моделей: {len(model_groups)}")
print(f"Моделей с несколькими вариантами: {len(multi_variant)}")

# Показываем первые 40 моделей с вариантами
print("\n\n=== МОДЕЛИ С НЕСКОЛЬКИМИ ВАРИАНТАМИ (первые 40) ===")
count = 0
for base, variants in sorted(multi_variant.items()):
    if count >= 40:
        break
    print(f"\n📱 {base}:")
    for v in variants:
        print(f"   {v['type']:12s} — {v['full_name']} — {v['price']} BYN")
    count += 1

# Подсчитаем типы
type_counts = defaultdict(int)
for name, price in all_data:
    dtype = extract_display_type(name)
    type_counts[dtype] += 1

print("\n\n=== РАСПРЕДЕЛЕНИЕ ТИПОВ ===")
for dtype, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
    print(f"  {dtype}: {cnt}")

wb.close()
