"""
Перестройка compatibility_display.json — все варианты для каждой модели
"""
import openpyxl
import json
import re
from collections import defaultdict

wb = openpyxl.load_workbook(r"C:\Users\user\Desktop\Дисплеи.xlsx", data_only=True)
ws = wb.active

# Собираем все данные
all_data = []
for row in range(1, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value or ""
    price = ws.cell(row=row, column=2).value
    all_data.append((name, price))

# Определяем тип дисплея
def extract_display_type(name):
    name_upper = name.upper()
    if 'IN-CELL' in name_upper:
        return 'In-Cell'
    if 'AMOLED' in name_upper:
        return 'AMOLED'
    if 'OLED' in name_upper:
        return 'OLED'
    if ' OR ' in name or '- OR' in name:
        return 'OR'
    if 'СТАНДАРТ' in name_upper:
        return 'Стандарт'
    return 'Стандарт'

# Убираем тип дисплея чтобы получить базовую модель
def get_base_model(name):
    model = name.replace("Дисплей ", "").strip()
    model = re.sub(r'\s*-\s*(OR|OLED|AMOLED|IN-CELL|LCD|TFT|Стандарт)\b', '', model, flags=re.IGNORECASE)
    model = re.sub(r'\s*\(.*?\)', '', model)
    return model.strip()

# Группируем по базовой модели телефона
model_groups = defaultdict(list)
for name, price in all_data:
    base = get_base_model(name)
    dtype = extract_display_type(name)
    model_groups[base].append({
        'full_name': name,
        'price': price,
        'type': dtype
    })

# Создаём новую структуру compatibility
# Формат: "phone_model" → список вариантов с типом, ценой и совместимостью
compatibility = {}
search_index = {}

for base_model, variants in model_groups.items():
    # Разбиваем base_model по слэшу — это совместимые модели
    phone_models = [m.strip() for m in base_model.split("/") if m.strip()]
    
    # Ключ группы — первая модель
    group_key = phone_models[0].lower().replace(" ", "_") + "_display_group"
    
    # Собираем все варианты
    options = []
    for v in variants:
        options.append({
            "type": v["type"],
            "price": v["price"],
            "full_name": v["full_name"]
        })
    
    compatibility[group_key] = {
        "models": phone_models,
        "options": options
    }
    
    # Строим поисковый индекс — каждая модель телефона → group_key
    for pm in phone_models:
        pm_lower = pm.lower()
        if pm_lower not in search_index:
            search_index[pm_lower] = []
        search_index[pm_lower].append(group_key)
        
        # Алиасы — подстроки
        words = pm_lower.split()
        for i in range(len(words)):
            for j in range(i+1, len(words)+1):
                alias = " ".join(words[i:j])
                if alias not in search_index:
                    search_index[alias] = []
                if group_key not in search_index[alias]:
                    search_index[alias].append(group_key)

# Сохраняем
output = {
    "compatibility": compatibility,
    "search_index": search_index
}

output_path = r"C:\Users\user\Desktop\Бот по стеклам\compatibility_display.json"
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

print(f"✅ Сохранено в {output_path}")
print(f"Групп: {len(compatibility)}")
print(f"Записей в поисковом индексе: {len(search_index)}")

# Считаем сколько моделей имеют несколько вариантов
multi_option = {k: v for k, v in compatibility.items() if len(v["options"]) > 1}
print(f"Моделей с несколькими вариантами: {len(multi_option)}")

# Примеры
print("\n=== ПРИМЕРЫ (первые 10 моделей с несколькими вариантами) ===")
count = 0
for key, data in multi_option.items():
    if count >= 10:
        break
    print(f"\n📱 {key}:")
    print(f"   Модели: {', '.join(data['models'])}")
    for opt in data['options']:
        print(f"   {opt['type']:12s} — {opt['full_name']} — {opt['price']} BYN")
    count += 1

wb.close()
