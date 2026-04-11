"""
Сравнение: Excel vs compatibility_display.json
"""
import openpyxl
import json
import re
from collections import defaultdict

# 1. Считаем записи в Excel
wb = openpyxl.load_workbook(r"C:\Users\user\Desktop\Дисплеи.xlsx", data_only=True)
ws = wb.active
excel_count = ws.max_row
print(f"📊 Excel записей: {excel_count}")

# Собираем все данные из Excel
excel_models = []
for row in range(1, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value or ""
    price = ws.cell(row=row, column=2).value
    excel_models.append((name, price))

# Определяем тип и базовую модель
def extract_display_type(name):
    name_upper = name.upper()
    if 'IN-CELL' in name_upper: return 'In-Cell'
    if 'AMOLED' in name_upper: return 'AMOLED'
    if 'OLED' in name_upper: return 'OLED'
    if ' OR ' in name or '- OR' in name: return 'OR'
    return 'Стандарт'

def get_base_model(name):
    model = name.replace("Дисплей ", "").strip()
    model = re.sub(r'\s*-\s*(OR|OLED|AMOLED|IN-CELL|LCD|TFT|Стандарт)\b', '', model, flags=re.IGNORECASE)
    model = re.sub(r'\s*\(.*?\)', '', model)
    return model.strip()

# Группируем Excel по базовой модели
excel_groups = defaultdict(list)
for name, price in excel_models:
    base = get_base_model(name)
    dtype = extract_display_type(name)
    excel_groups[base].append({'type': dtype, 'price': price, 'full_name': name})

# 2. Считаем записи в JSON
with open(r"C:\Users\user\Desktop\Бот по стеклам\compatibility_display.json", "r", encoding="utf-8") as f:
    json_data = json.load(f)

compatibility = json_data["compatibility"]
json_total_options = sum(len(g["options"]) for g in compatibility.values())
json_groups = len(compatibility)

print(f"📊 JSON групп: {json_groups}")
print(f"📊 JSON вариантов (options): {json_total_options}")
print(f"📊 Excel групп (уникальных моделей): {len(excel_groups)}")

# Ищем расхождения
if excel_count == json_total_options:
    print(f"\n✅ ВСЕ {excel_count} записей из Excel в JSON!")
else:
    print(f"\n⚠️ РАСХОЖДЕНИЕ: Excel={excel_count}, JSON options={json_total_options}")
    diff = excel_count - json_total_options
    if diff > 0:
        print(f"   НЕ ХВАТАЕТ {diff} записей в JSON!")
    else:
        print(f"   В JSON на {-diff} записей БОЛЬШЕ")

# Проверяем что каждая модель из Excel есть в JSON
missing_in_json = []
for base_model, variants in excel_groups.items():
    # Ищем в JSON
    found = False
    for group_key, group_data in compatibility.items():
        json_models_lower = [m.lower() for m in group_data["models"]]
        # Разбиваем base_model по слэшу
        excel_models_split = [m.strip().lower() for m in base_model.split("/")]
        if any(em in json_models_lower for em in excel_models_split):
            found = True
            break
    if not found:
        missing_in_json.append(base_model)

if missing_in_json:
    print(f"\n❌ Моделей из Excel НЕТ в JSON ({len(missing_in_json)}):")
    for m in missing_in_json[:20]:
        print(f"   - {m}")
    if len(missing_in_json) > 20:
        print(f"   ... и ещё {len(missing_in_json) - 20}")
else:
    print(f"\n✅ ВСЕ модели из Excel найдены в JSON!")

# Проверяем что ВСЕ варианты для каждой модели
print(f"\n=== ПРОВЕРКА ПОЛНОТЫ ВАРИАНТОВ ===")
incomplete_models = []
for base_model, excel_variants in excel_groups.items():
    excel_types = set(v['type'] for v in excel_variants)
    
    # Ищем в JSON
    for group_key, group_data in compatibility.items():
        json_models_lower = [m.lower() for m in group_data["models"]]
        excel_models_split = [m.strip().lower() for m in base_model.split("/")]
        if any(em in json_models_lower for em in excel_models_split):
            json_types = set(opt['type'] for opt in group_data["options"])
            if excel_types != json_types:
                incomplete_models.append({
                    'model': base_model,
                    'excel_types': excel_types,
                    'json_types': json_types,
                    'excel_count': len(excel_variants),
                    'json_count': len(group_data["options"])
                })
            break

if incomplete_models:
    print(f"\n⚠️ Моделей с неполными вариантами: {len(incomplete_models)}")
    for item in incomplete_models[:15]:
        print(f"\n   📱 {item['model']}:")
        print(f"      Excel: {item['excel_types']} ({item['excel_count']} шт)")
        print(f"      JSON:  {item['json_types']} ({item['json_count']} шт)")
else:
    print(f"\n✅ ВСЕ варианты для ВСЕХ моделей добавлены!")

wb.close()
