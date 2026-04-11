"""
Перестройка compatibility_display.json — ВСЕ 1450 записей из Excel без потерь

Подход:
1. Каждая строка из Excel = один вариант дисплея
2. Группировка по НОРМАЛИЗОВАННОЙ модели телефона (без типа дисплея)
3. Каждая запись сохраняется
"""
import openpyxl
import json
import re
from collections import defaultdict, OrderedDict

wb = openpyxl.load_workbook(r"C:\Users\user\Desktop\Дисплеи.xlsx", data_only=True)
ws = wb.active

# Определяем тип дисплея
def extract_display_type(name):
    name_upper = name.upper()
    if 'IN-CELL' in name_upper: return 'In-Cell'
    if 'AMOLED' in name_upper: return 'AMOLED'
    if 'OLED' in name_upper: return 'OLED'
    if ' OR ' in name or '- OR' in name: return 'OR'
    if 'СТАНДАРТ' in name_upper: return 'Стандарт'
    return 'Стандарт'

# Нормализуем название — убираем тип дисплея, но сохраняем модели телефонов
def normalize_phone_models(name):
    """Извлекаем названия моделей телефонов из строки"""
    # Убираем "Дисплей "
    model = name.replace("Дисплей ", "").strip()
    # Убираем тип дисплея
    model = re.sub(r'\s*-\s*(In-Cell|OR|OLED|AMOLED|LCD|TFT|Стандарт)\b', '', model, flags=re.IGNORECASE).strip()
    # Убираем примечания в скобках
    model = re.sub(r'\s*\(.*?\)', '', model).strip()
    return model

# Группируем по нормализованной модели
# Ключ = нормализованная строка моделей
model_groups = defaultdict(list)

for row in range(1, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value or ""
    price = ws.cell(row=row, column=2).value
    
    dtype = extract_display_type(name)
    normalized = normalize_phone_models(name)
    
    # Сохраняем ВСЕ данные
    model_groups[normalized].append({
        'type': dtype,
        'price': price,
        'full_name': name,
        'normalized': normalized
    })

print(f"Excel записей: {ws.max_row}")
print(f"JSON групп (после нормализации): {len(model_groups)}")

# Считаем сколько записей в итоге
total_options = sum(len(v) for v in model_groups.values())
print(f"JSON вариантов (options): {total_options}")

if total_options == ws.max_row:
    print("✅ ВСЕ записи сохранены!")
else:
    print(f"⚠️ ПОТЕРЯНО {ws.max_row - total_options} записей!")

# Создаём структуру
compatibility = OrderedDict()
search_index = {}

group_counter = 0
for normalized_models_str, variants in model_groups.items():
    group_counter += 1
    
    # Разбиваем по слэшу — это совместимые модели телефонов
    phone_models = [m.strip() for m in normalized_models_str.split("/") if m.strip()]
    
    # Ключ группы
    group_key = f"display_{group_counter:04d}"
    
    # Собираем варианты — объединяем одинаковые типы с разными рамами
    by_type = defaultdict(list)
    for v in variants:
        by_type[v['type']].append(v)
    
    unique_options = []
    for dtype, items in by_type.items():
        if len(items) == 1:
            unique_options.append({
                "type": items[0]["type"],
                "price": items[0]["price"],
                "full_name": items[0]["full_name"]
            })
        else:
            # Несколько одинаковых типов — берём диапазон цен
            prices = [i['price'] for i in items if i['price']]
            min_price = min(prices) if prices else 0
            max_price = max(prices) if prices else 0
            # Берём самый дорогой как основной (обычно полная версия)
            best = max(items, key=lambda x: x["price"] if x["price"] else 0)
            note = f"Цена: {min_price}-{max_price} BYN (зависит от рамы)"
            unique_options.append({
                "type": best["type"],
                "price": max_price,
                "full_name": best["full_name"],
                "note": note
            })
    
    compatibility[group_key] = {
        "models": phone_models,
        "options": unique_options
    }
    
    # Поисковый индекс — каждая модель телефона → group_key
    for pm in phone_models:
        pm_lower = pm.lower()
        if pm_lower not in search_index:
            search_index[pm_lower] = []
        if group_key not in search_index[pm_lower]:
            search_index[pm_lower].append(group_key)
        
        # Алиасы
        words = pm_lower.split()
        for i in range(len(words)):
            for j in range(i+1, len(words)+1):
                alias = " ".join(words[i:j])
                if alias not in search_index:
                    search_index[alias] = []
                if group_key not in search_index[alias]:
                    search_index[alias].append(group_key)

# Сохраняем
output = {
    "compatibility": compatibility,
    "search_index": search_index
}

output_path = r"C:\Users\user\Desktop\Бот по стеклам\compatibility_display.json"
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

print(f"\n✅ Сохранено в {output_path}")
print(f"Групп: {len(compatibility)}")
print(f"Записей в поисковом индексе: {len(search_index)}")

# Статистика по типам
type_counts = defaultdict(int)
for gdata in compatibility.values():
    for opt in gdata["options"]:
        type_counts[opt["type"]] += 1

print(f"\nРаспределение типов:")
for t, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
    print(f"  {t}: {cnt}")

# Модели с несколькими вариантами
multi = {k: v for k, v in compatibility.items() if len(v["options"]) > 1}
print(f"\nМоделей с несколькими вариантами: {len(multi)}")

# Проверка — первые 10 моделей с несколькими вариантами
print(f"\n=== ПРИМЕРЫ (первые 15) ===")
count = 0
for key, gdata in multi.items():
    if count >= 15:
        break
    print(f"\n📱 {key}:")
    print(f"   Модели: {', '.join(gdata['models'])}")
    for opt in gdata['options']:
        print(f"   {opt['type']:12s} — {opt['price']} BYN")
    count += 1

wb.close()
